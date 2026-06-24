"""Export + import helpers for PLUMBLINE — Excel & PDF recap, AI-assisted xlsx import."""
import io
import json
import logging
from typing import List, Dict, Any
from datetime import datetime, timezone

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_LEFT


# ── COLORS (matching PLUMBLINE brand) ──────────────────────────────
ORANGE = "FF5F15"
YELLOW = "CCFF00"
DARK = "09090B"
DARK_SURF = "18181B"
GREY = "A1A1AA"
WHITE = "FAFAFA"


def _h_fill(color: str):
    return PatternFill("solid", fgColor=color)


def build_recap_xlsx(job: dict, tasks: list, entries: list, dashboard: dict) -> bytes:
    """Build a styled multi-sheet Excel recap."""
    wb = Workbook()

    # ─── Sheet 1: Executive Summary ────────────────────────────────
    s = wb.active
    s.title = "Recap"

    # Title block
    s["A1"] = "PLUMBLINE — JOB RECAP"
    s["A1"].font = Font(name="Calibri", size=22, bold=True, color=ORANGE)
    s.merge_cells("A1:F1")
    s["A2"] = job.get("name", "")
    s["A2"].font = Font(name="Calibri", size=16, bold=True)
    s.merge_cells("A2:F2")
    s["A3"] = f"{job.get('location','')} · {job.get('client','')} · Status: {job.get('status','active').upper()}"
    s["A3"].font = Font(name="Calibri", size=11, color="666666")
    s.merge_cells("A3:F3")
    s["A4"] = f"Generated {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    s["A4"].font = Font(name="Calibri", size=9, italic=True, color="999999")
    s.merge_cells("A4:F4")

    # KPI grid
    totals = dashboard.get("totals", {})
    roi = dashboard.get("roi", {})
    vstats = dashboard.get("validation_stats", {})
    kpi_row = 6
    kpis = [
        ("Estimated Hrs", totals.get("estimated_hours", 0)),
        ("Actual Hrs", totals.get("actual_hours", 0)),
        ("Earned Hrs", totals.get("earned_hours", 0)),
        ("Variance", totals.get("variance_hours", 0)),
        ("Ratio", totals.get("ratio", 0)),
        ("$ Protected", f"${(roi.get('total_value_protected', 0)):,.0f}"),
    ]
    for i, (label, value) in enumerate(kpis):
        col = get_column_letter(i + 1)
        s[f"{col}{kpi_row}"] = label
        s[f"{col}{kpi_row}"].font = Font(name="Calibri", size=9, bold=True, color="666666")
        s[f"{col}{kpi_row+1}"] = value
        s[f"{col}{kpi_row+1}"].font = Font(name="Calibri", size=18, bold=True)
        s[f"{col}{kpi_row+1}"].fill = _h_fill("F4F4F5")
        s[f"{col}{kpi_row+1}"].alignment = Alignment(horizontal="center")

    # Validation Layer summary
    base = kpi_row + 4
    s[f"A{base}"] = "VALIDATION LAYER"
    s[f"A{base}"].font = Font(name="Calibri", size=14, bold=True, color=ORANGE)
    s.merge_cells(f"A{base}:F{base}")
    val_cells = [
        ("Checks Run", vstats.get("total", 0)),
        ("Pass Rate", f"{int(vstats.get('pass_rate', 0) * 100)}%"),
        ("Failed (caught)", vstats.get("failed", 0)),
        ("Photos Logged", vstats.get("photos_captured", 0)),
    ]
    for i, (k, v) in enumerate(val_cells):
        s[f"{get_column_letter(i+1)}{base+1}"] = k
        s[f"{get_column_letter(i+1)}{base+1}"].font = Font(bold=True, size=10, color="666666")
        s[f"{get_column_letter(i+1)}{base+2}"] = v
        s[f"{get_column_letter(i+1)}{base+2}"].font = Font(size=14, bold=True)

    # Auto-width columns
    for col in range(1, 7):
        s.column_dimensions[get_column_letter(col)].width = 22

    # ─── Sheet 2: Task Detail ──────────────────────────────────────
    s2 = wb.create_sheet("Tasks")
    headers = ["Category", "Course", "Task", "Unit", "Est Hrs", "Actual Hrs", "Est Qty", "Actual Qty", "Status"]
    for i, h in enumerate(headers, 1):
        c = s2.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = _h_fill("18181B")
        c.alignment = Alignment(horizontal="left", vertical="center")
    for r, t in enumerate(tasks, 2):
        s2.cell(row=r, column=1, value=t.get("category"))
        s2.cell(row=r, column=2, value=t.get("course"))
        s2.cell(row=r, column=3, value=t.get("name"))
        s2.cell(row=r, column=4, value=t.get("unit") or "")
        s2.cell(row=r, column=5, value=t.get("estimated_hours") or 0)
        s2.cell(row=r, column=6, value=t.get("actual_hours") or 0)
        s2.cell(row=r, column=7, value=t.get("estimated_qty") or 0)
        s2.cell(row=r, column=8, value=t.get("actual_qty") or 0)
        status_cell = s2.cell(row=r, column=9, value=t.get("status", "").replace("_", " ").upper())
        if t.get("status") == "validated":
            status_cell.fill = _h_fill("D9F99D")
        elif t.get("status") == "rework":
            status_cell.fill = _h_fill("FED7AA")
        elif t.get("status") == "in_progress":
            status_cell.fill = _h_fill("BFDBFE")
    widths = [12, 8, 50, 8, 10, 10, 10, 10, 14]
    for i, w in enumerate(widths, 1):
        s2.column_dimensions[get_column_letter(i)].width = w
    s2.freeze_panes = "A2"

    # ─── Sheet 3: Crew Leaderboard ─────────────────────────────────
    s3 = wb.create_sheet("Leaderboard")
    headers = ["Rank", "Name", "Role", "Hours", "Entries", "Checks", "Pass Rate", "Catches", "Photos", "Score"]
    for i, h in enumerate(headers, 1):
        c = s3.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = _h_fill("18181B")
    for r, p in enumerate(dashboard.get("leaderboard", []), 2):
        s3.cell(row=r, column=1, value=r - 1)
        s3.cell(row=r, column=2, value=p.get("name"))
        s3.cell(row=r, column=3, value=p.get("role"))
        s3.cell(row=r, column=4, value=p.get("hours"))
        s3.cell(row=r, column=5, value=p.get("entries"))
        s3.cell(row=r, column=6, value=p.get("checks_total"))
        s3.cell(row=r, column=7, value=f"{int(p.get('pass_rate', 0) * 100)}%")
        s3.cell(row=r, column=8, value=p.get("checks_failed"))
        s3.cell(row=r, column=9, value=p.get("photos"))
        s3.cell(row=r, column=10, value=p.get("score"))
    for i, w in enumerate([6, 22, 14, 10, 10, 10, 12, 10, 10, 10], 1):
        s3.column_dimensions[get_column_letter(i)].width = w
    s3.freeze_panes = "A2"

    # ─── Sheet 4: Entries (last 200) ──────────────────────────────
    s4 = wb.create_sheet("Entries")
    headers = ["Date", "Task", "Crew", "Role", "Hours", "Qty", "Failed?", "Notes"]
    for i, h in enumerate(headers, 1):
        c = s4.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = _h_fill("18181B")
    # Build task_id -> name map
    task_name_map = {t["id"]: t["name"] for t in tasks}
    for r, e in enumerate(entries[:200], 2):
        s4.cell(row=r, column=1, value=e.get("created_at", "")[:10])
        s4.cell(row=r, column=2, value=task_name_map.get(e.get("task_id"), e.get("task_id")))
        s4.cell(row=r, column=3, value=e.get("crew_member"))
        s4.cell(row=r, column=4, value=e.get("role"))
        s4.cell(row=r, column=5, value=e.get("hours"))
        s4.cell(row=r, column=6, value=e.get("qty_completed"))
        s4.cell(row=r, column=7, value="YES" if e.get("has_failed_check") else "")
        s4.cell(row=r, column=8, value=(e.get("notes") or "")[:200])
    for i, w in enumerate([12, 40, 18, 14, 10, 10, 10, 50], 1):
        s4.column_dimensions[get_column_letter(i)].width = w
    s4.freeze_panes = "A2"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def build_recap_pdf(job: dict, tasks: list, entries: list, dashboard: dict) -> bytes:
    """Build a clean executive KPI PDF report."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, leftMargin=40, rightMargin=40, topMargin=40, bottomMargin=40)
    elements = []
    styles = getSampleStyleSheet()

    # Custom styles
    h_brand = ParagraphStyle("brand", parent=styles["Title"], fontName="Helvetica-Bold",
                             fontSize=28, textColor=colors.HexColor("#FF5F15"), alignment=TA_LEFT, spaceAfter=4)
    h_job = ParagraphStyle("job", parent=styles["Heading2"], fontName="Helvetica-Bold",
                           fontSize=18, textColor=colors.HexColor("#18181B"), spaceAfter=2)
    h_meta = ParagraphStyle("meta", parent=styles["Normal"], fontName="Helvetica",
                            fontSize=9, textColor=colors.HexColor("#71717A"), spaceAfter=12)
    h_sec = ParagraphStyle("sec", parent=styles["Heading3"], fontName="Helvetica-Bold",
                           fontSize=13, textColor=colors.HexColor("#FF5F15"),
                           spaceBefore=14, spaceAfter=6)

    elements.append(Paragraph("PLUMBLINE", h_brand))
    elements.append(Paragraph(job.get("name", ""), h_job))
    elements.append(Paragraph(
        f"{job.get('location','')} &nbsp;·&nbsp; {job.get('client','')} &nbsp;·&nbsp; "
        f"Status: <b>{job.get('status','active').upper()}</b> &nbsp;·&nbsp; "
        f"Generated {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
        h_meta,
    ))

    # ─── KPI tiles ─────────────────────────────────────────────────
    totals = dashboard.get("totals", {})
    roi = dashboard.get("roi", {})
    vstats = dashboard.get("validation_stats", {})

    elements.append(Paragraph("Key Metrics", h_sec))
    kpi_rows = [
        ["Estimated Hrs", "Actual Hrs", "Earned Hrs", "Variance", "Ratio", "$ Protected"],
        [
            f"{totals.get('estimated_hours', 0):.1f}",
            f"{totals.get('actual_hours', 0):.1f}",
            f"{totals.get('earned_hours', 0):.1f}",
            f"{totals.get('variance_hours', 0):.1f}",
            f"{totals.get('ratio', 0):.2f}",
            f"${roi.get('total_value_protected', 0):,.0f}",
        ],
    ]
    kpi_t = Table(kpi_rows, colWidths=[1.2*inch]*6)
    kpi_t.setStyle(TableStyle([
        ("FONT", (0, 0), (-1, 0), "Helvetica-Bold", 8),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#71717A")),
        ("FONT", (0, 1), (-1, 1), "Helvetica-Bold", 16),
        ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#F4F4F5")),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#E4E4E7")),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E4E4E7")),
        ("BOTTOMPADDING", (0, 1), (-1, 1), 12),
        ("TOPPADDING", (0, 1), (-1, 1), 12),
    ]))
    elements.append(kpi_t)

    # ─── Validation Layer ──────────────────────────────────────────
    elements.append(Paragraph("Validation Layer", h_sec))
    val_rows = [
        ["Checks Run", "Pass Rate", "Failed Caught", "Photos Logged"],
        [
            str(vstats.get("total", 0)),
            f"{int(vstats.get('pass_rate', 0) * 100)}%",
            str(vstats.get("failed", 0)),
            str(vstats.get("photos_captured", 0)),
        ],
    ]
    val_t = Table(val_rows, colWidths=[1.7*inch]*4)
    val_t.setStyle(TableStyle([
        ("FONT", (0, 0), (-1, 0), "Helvetica-Bold", 9),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#71717A")),
        ("FONT", (0, 1), (-1, 1), "Helvetica-Bold", 18),
        ("TEXTCOLOR", (1, 1), (1, 1), colors.HexColor("#65A30D")),
        ("TEXTCOLOR", (2, 1), (2, 1), colors.HexColor("#FF5F15")),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#E4E4E7")),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E4E4E7")),
        ("BOTTOMPADDING", (0, 1), (-1, 1), 10),
        ("TOPPADDING", (0, 1), (-1, 1), 10),
    ]))
    elements.append(val_t)

    # ─── Leaderboard ──────────────────────────────────────────────
    leaderboard = dashboard.get("leaderboard", [])
    if leaderboard:
        elements.append(Paragraph("Foreman Leaderboard", h_sec))
        lb_rows = [["#", "Name", "Role", "Hrs", "Pass %", "Catches", "Score"]]
        for i, p in enumerate(leaderboard[:10], 1):
            lb_rows.append([
                str(i), p.get("name", ""), p.get("role", ""),
                f"{p.get('hours', 0):.1f}",
                f"{int(p.get('pass_rate', 0) * 100)}%",
                str(p.get("checks_failed", 0)),
                f"{p.get('score', 0):.1f}",
            ])
        lb_t = Table(lb_rows, colWidths=[0.4*inch, 2.2*inch, 1.0*inch, 0.7*inch, 0.8*inch, 0.8*inch, 0.7*inch])
        lb_t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#18181B")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#FAFAFA")),
            ("FONT", (0, 0), (-1, 0), "Helvetica-Bold", 9),
            ("FONT", (0, 1), (-1, -1), "Helvetica", 9),
            ("ALIGN", (3, 0), (-1, -1), "RIGHT"),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#E4E4E7")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(lb_t)

    # ─── Task summary by category ──────────────────────────────────
    elements.append(Paragraph("Tasks by Phase", h_sec))
    cat_breakdown = dashboard.get("category_breakdown", {})
    cat_rows = [["Phase", "Total", "Validated", "Rework", "Est Hrs", "Actual Hrs"]]
    for cat, d in cat_breakdown.items():
        cat_rows.append([cat, d.get("total", 0), d.get("validated", 0), d.get("rework", 0),
                         f"{d.get('est_hours', 0):.1f}", f"{d.get('actual_hours', 0):.1f}"])
    cat_t = Table(cat_rows, colWidths=[1.4*inch, 0.8*inch, 0.9*inch, 0.8*inch, 1.0*inch, 1.0*inch])
    cat_t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#18181B")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#FAFAFA")),
        ("FONT", (0, 0), (-1, 0), "Helvetica-Bold", 9),
        ("FONT", (0, 1), (-1, -1), "Helvetica", 9),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#E4E4E7")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(cat_t)

    # ─── Rework hotlist ────────────────────────────────────────────
    rework_tasks = dashboard.get("rework_tasks", [])
    if rework_tasks:
        elements.append(Paragraph("Active Rework Flags", h_sec))
        rw_rows = [["Task", "Phase", "Course"]]
        for t in rework_tasks[:15]:
            rw_rows.append([t.get("name", ""), t.get("category", ""), t.get("course", "")])
        rw_t = Table(rw_rows, colWidths=[4.2*inch, 1.2*inch, 1.0*inch])
        rw_t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FF5F15")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONT", (0, 0), (-1, 0), "Helvetica-Bold", 9),
            ("FONT", (0, 1), (-1, -1), "Helvetica", 9),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#E4E4E7")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FFF7ED")]),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(rw_t)

    doc.build(elements)
    return buf.getvalue()


def parse_xlsx_for_import(file_bytes: bytes, max_rows: int = 400) -> List[Dict[str, Any]]:
    """Read an xlsx file and return rows of {row_idx, sheet, columns: {header: value}}.
    Tries to detect a header row, then yields all data rows below it.
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    out = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        # Find header row — first row with >=3 non-empty cells, prefer earlier
        header_idx = 0
        for idx, row in enumerate(rows[:20]):
            non_empty = sum(1 for c in row if c not in (None, ""))
            if non_empty >= 3:
                header_idx = idx
                break
        headers = [str(c).strip() if c else f"col_{i}" for i, c in enumerate(rows[header_idx])]
        for r_idx, row in enumerate(rows[header_idx + 1:][:max_rows], start=header_idx + 2):
            non_empty = sum(1 for c in row if c not in (None, ""))
            if non_empty < 1:
                continue
            entry = {"sheet": sheet_name, "row": r_idx, "columns": {}}
            for h, v in zip(headers, row):
                if v not in (None, ""):
                    entry["columns"][h] = v
            if entry["columns"]:
                out.append(entry)
    return out
